#!/usr/bin/env python3
"""
Merged pipeline:
1) Download a Google Sheet (by spreadsheet ID) into leads_agent_excel_files/leads_ddmmyyyy.xlsx
2) Read that XLSX and generate a report (TXT) exactly like your current flow,
   BUT only feed rows to the model where bounce_reason is NOT None/NaN/"" (after strip).
3) (Optional) If you store Gmail message IDs, you can also pull those emails and include snippets.

Notes:
- Uses Google Sheets API (values -> xlsx) like the earlier script.
- Keeps your llama_cpp + report formatting logic.
"""
import os
import sys
import json
import signal
import datetime as dt
from pathlib import Path
from datetime import datetime, timedelta

import pandas as pd
import pytz
from openpyxl import Workbook

from llama_cpp import Llama

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build

# -----------------------------
# CONFIG - Load from JSON
# -----------------------------
def load_config(config_file: str = "config.json") -> dict:
    """Load configuration from JSON file."""
    with open(config_file, "r", encoding="utf-8") as f:
        return json.load(f)

config = load_config()

MODEL_PATH = config["MODEL_PATH"]
SCOPES = config["SCOPES"]
CREDENTIALS_JSON = config["CREDENTIALS_JSON"]
TOKEN_JSON = config["TOKEN_JSON"]
SPREADSHEET_ID = config["SPREADSHEET_ID"]
OUTPUT_DIR = Path(config["OUTPUT_DIR"])
OUTPUT_PREFIX = config["OUTPUT_PREFIX"]
COL_SENT_AT = config["COL_SENT_AT"]
COL_VERIFIED_AT = config["COL_VERIFIED_AT"]
COL_BOUNCE_REASON = config["COL_BOUNCE_REASON"]
ENABLE_GMAIL_PULL = config["ENABLE_GMAIL_PULL"]
GMAIL_SCOPES = config["GMAIL_SCOPES"]
COL_GMAIL_MSG_ID = config["COL_GMAIL_MSG_ID"]
MAX_GMAIL_BODY_CHARS = config["MAX_GMAIL_BODY_CHARS"]
# -----------------------------


def get_creds(scopes, credentials_path=CREDENTIALS_JSON, token_path=TOKEN_JSON) -> Credentials:
    creds = None
    if os.path.exists(token_path):
        creds = Credentials.from_authorized_user_file(token_path, scopes)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not os.path.exists(credentials_path):
                raise FileNotFoundError(f"Missing {credentials_path}")
            flow = InstalledAppFlow.from_client_secrets_file(credentials_path, scopes)
            creds = flow.run_local_server(port=0)
        with open(token_path, "w", encoding="utf-8") as f:
            f.write(creds.to_json())

    return creds


def download_google_sheet_to_xlsx(spreadsheet_id: str, out_path: Path) -> None:
    """
    Pull all tabs via Sheets API (values) and write to a local .xlsx.
    """
    creds = get_creds(SCOPES)
    service = build("sheets", "v4", credentials=creds)

    meta = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
    sheets = meta.get("sheets", [])
    if not sheets:
        raise RuntimeError("Spreadsheet has no tabs.")

    wb = Workbook()
    ws0 = wb.active
    ws0.title = sheets[0]["properties"]["title"]

    for i, sh in enumerate(sheets):
        title = sh["properties"]["title"]
        ws = ws0 if i == 0 else wb.create_sheet(title=title)

        resp = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range=title
        ).execute()
        values = resp.get("values", [])

        for r_idx, row in enumerate(values, start=1):
            for c_idx, cell in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=cell)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def try_fetch_gmail_message_text(gmail_message_id: str) -> str:
    """
    Optional: fetch a Gmail message plain text snippet/body.
    Requires token authorized for gmail.readonly scope.
    We keep it short to avoid blowing context.
    """
    if not gmail_message_id or str(gmail_message_id).strip() == "":
        return ""

    creds = get_creds(GMAIL_SCOPES)
    gmail = build("gmail", "v1", credentials=creds)
    msg = gmail.users().messages().get(userId="me", id=str(gmail_message_id), format="full").execute()

    snippet = msg.get("snippet", "") or ""
    # Full MIME decode is more involved; snippet is usually enough for your use-case.
    text = snippet.strip()
    if len(text) > MAX_GMAIL_BODY_CHARS:
        text = text[:MAX_GMAIL_BODY_CHARS] + "…"
    return text


class SummaryBot:
    def __init__(self, model_path: str, n_ctx: int = 16384):
        if not os.path.exists(model_path):
            raise FileNotFoundError(f"Model not found: {os.path.abspath(model_path)}")

        self.history = [
            {"role": "system", "content": "You are a precise assistant specializing in summarizing email lead status and reasons for no response."}
        ]

        self.llm = Llama(
            model_path=model_path,
            n_gpu_layers=-1,
            n_ctx=n_ctx,
            n_threads=12,
            flash_attn=True,
            verbose=False
        )

    def chat(self, user_query: str):
        self.history.append({"role": "user", "content": user_query})
        response_stream = self.llm.create_chat_completion(
            messages=self.history,
            stream=True,
            temperature=0.2,
            max_tokens=2048
        )
        full_response = ""
        for chunk in response_stream:
            delta = chunk["choices"][0].get("delta", {})
            if "content" in delta:
                content = delta["content"]
                full_response += content
                yield content
        self.history.append({"role": "assistant", "content": full_response})


class ReportGenerator:
    def __init__(self, model_path: str):
        self.model_path = model_path

    @staticmethod
    def format_text_with_line_breaks(text: str, words_per_line: int = 15) -> str:
        words = (text or "").split()
        lines = []
        for i in range(0, len(words), words_per_line):
            lines.append(" ".join(words[i:i + words_per_line]))
        return "\n".join(lines)

    @staticmethod
    def _is_nonempty_bounce_reason(x) -> bool:
        """
        The requested change:
        keep only rows where bounce_reason is NOT None, NOT NaN, NOT "" (after strip).
        """
        if x is None:
            return False
        # pandas NaN
        try:
            if pd.isna(x):
                return False
        except Exception:
            pass
        s = str(x).strip()
        return s != "" and s.lower() not in {"none", "null"}

    def generate_report_from_xlsx(self, xlsx_path: Path, report_id: str) -> Path:
        if not xlsx_path.exists():
            raise FileNotFoundError(f"Excel file not found: {xlsx_path}")

        df = pd.read_excel(xlsx_path)

        # Time window
        ist = pytz.timezone("Asia/Kolkata")
        current_time = datetime.now(ist)
        last_24hrs = current_time - timedelta(hours=24)

        # Parse timestamps
        df[COL_SENT_AT] = pd.to_datetime(df[COL_SENT_AT], errors="coerce")
        # verified_at can be blank; keep as datetime where possible
        if COL_VERIFIED_AT in df.columns:
            df[COL_VERIFIED_AT] = pd.to_datetime(df[COL_VERIFIED_AT], errors="coerce")

        # Filter: last 24h + verified
        filtered = df[df[COL_SENT_AT] > last_24hrs]
        if COL_VERIFIED_AT in filtered.columns:
            filtered = filtered[filtered[COL_VERIFIED_AT].notnull()]

        # NEW FILTER: only rows with bounce_reason present
        if COL_BOUNCE_REASON in filtered.columns:
            filtered = filtered[filtered[COL_BOUNCE_REASON].apply(self._is_nonempty_bounce_reason)]
        else:
            # If column missing, nothing to summarize under your new constraint
            filtered = filtered.iloc[0:0]

        # Build report
        report_content = ""
        report_content += "=" * 117 + "\n"
        report_content += "OFFICIAL EMAIL LEADS STATUS REPORT\n"
        report_content += "=" * 117 + "\n"
        report_content += f"REPORT ID: {report_id}\n"
        report_content += f"GENERATED ON: {current_time.strftime('%Y-%m-%d %H:%M:%S %Z')}\n"
        report_content += (
            f"PERIOD: Last 24 Hours (From {last_24hrs.strftime('%Y-%m-%d %H:%M:%S %Z')} "
            f"to {current_time.strftime('%Y-%m-%d %H:%M:%S %Z')})\n"
        )
        report_content += "=" * 117 + "\n\n"

        if filtered.empty:
            report_content += "No verified leads in the last 24 hours with a non-empty bounce_reason.\n"
        else:
            for _, row in filtered.iterrows():
                bot = SummaryBot(self.model_path)

                gmail_msg_id = row.get(COL_GMAIL_MSG_ID, "N/A")
                gmail_excerpt = ""
                if ENABLE_GMAIL_PULL and COL_GMAIL_MSG_ID in filtered.columns:
                    try:
                        gmail_excerpt = try_fetch_gmail_message_text(str(gmail_msg_id))
                    except Exception as e:
                        gmail_excerpt = f"(Gmail fetch failed: {e})"

                query = f"""Summarize in a professional manner why this lead did not respond, based on the following data:

Lead ID: {row.get('lead_id', 'N/A')}
Email: {row.get('email', 'N/A')}
First Name: {row.get('first_name', 'N/A')}
Company: {row.get('company', 'N/A')}
Status: {row.get('status', 'N/A')}
Sent At: {row.get(COL_SENT_AT, 'N/A')}
Gmail Msg ID: {gmail_msg_id}
Bounce Code: {row.get('bounce_code', 'N/A')}
Bounce Reason: {row.get(COL_BOUNCE_REASON, 'N/A')}
Verified At: {row.get(COL_VERIFIED_AT, 'N/A')}

Optional Gmail excerpt (if present):
{gmail_excerpt}

Provide a short summary focused on the reason for no response. Prefer concrete operational causes (delivery failure, policy blocks, invalid address, etc.) over speculation.
"""

                full_summary = ""
                for chunk in bot.chat(query):
                    full_summary += chunk

                formatted_summary = self.format_text_with_line_breaks(full_summary, words_per_line=15)

                report_content += "-" * 117 + "\n"
                report_content += f"LEAD ID: {row.get('lead_id', 'N/A')}\n"
                report_content += f"NAME: {row.get('first_name', 'N/A')}\n"
                report_content += f"COMPANY: {row.get('company', 'N/A')}\n"
                report_content += f"EMAIL: {row.get('email', 'N/A')}\n"
                report_content += f"STATUS: {row.get('status', 'N/A')}\n"
                report_content += "BOUNCE_REASON_SUMMARY:\n"
                report_content += formatted_summary + "\n"
                report_content += "-" * 117 + "\n\n"

        report_content += "=" * 117 + "\n"
        report_content += "END OF REPORT\n"
        report_content += "=" * 117 + "\n"

        report_dir = Path("excel_leads_daily_list")
        report_dir.mkdir(parents=True, exist_ok=True)
        report_path = report_dir / f"report_{report_id}.txt"
        with open(report_path, "w", encoding="utf-8") as f:
            f.write(report_content)

        return report_path


def main() -> int:
    # Download sheet to dated XLSX
    ddmmyyyy = dt.datetime.now().strftime("%d%m%Y")
    xlsx_path = OUTPUT_DIR / f"{OUTPUT_PREFIX}{ddmmyyyy}.xlsx"

    download_google_sheet_to_xlsx(SPREADSHEET_ID, xlsx_path)
    print(f"[OK] Saved XLSX: {xlsx_path.resolve()}")

    # Generate report from that XLSX
    model_abs = os.path.abspath(MODEL_PATH)
    gen = ReportGenerator(model_abs)
    report_path = gen.generate_report_from_xlsx(xlsx_path=xlsx_path, report_id=ddmmyyyy)
    print(f"[OK] Report generated: {report_path.resolve()}")
    return 0


if __name__ == "__main__":
    # Clean ctrl+c handling
    def _sigint(_sig, _frame):
        print("\n[STOP] Interrupted.", file=sys.stderr)
        raise SystemExit(130)

    signal.signal(signal.SIGINT, _sigint)

    raise SystemExit(main())
